import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Agents Productivity Daily Summary", layout="wide")

st.title("📊 Agents Productivity Daily Summary")
st.markdown("Upload multiple XLSX files → Daily Agent Productivity Report (Unique RPC by Debtor ID)")

# ========================= SIDEBAR =========================
with st.sidebar:
    st.header("📁 Data Upload")
    uploaded_files = st.file_uploader(
        "Upload one or more .xlsx files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="You can select and upload multiple files at once"
    )
    st.markdown("---")

if not uploaded_files:
    st.info("👈 Please upload your Excel files using the sidebar uploader.")
    st.stop()

# ========================= LOAD & COMBINE FILES =========================
with st.spinner("Processing files..."):
    dfs = []
    for uploaded_file in uploaded_files:
        try:
            temp_df = pd.read_excel(uploaded_file)
            dfs.append(temp_df)
        except Exception as e:
            st.error(f"❌ Error reading {uploaded_file.name}: {e}")
            st.stop()

    df = pd.concat(dfs, ignore_index=True)

# ========================= DATA PREPARATION =========================
df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date

numeric_cols = ['Talk Time Duration', 'PTP Amount', 'Balance', 'Claim Paid Amount']
for col in numeric_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

df['Status'] = df['Status'].fillna('').astype(str)
df['Remark By'] = df['Remark By'].fillna('Unknown').astype(str)
df['Debtor ID'] = df['Debtor ID'].fillna('').astype(str)   # Important for unique counting

# ========================= CALCULATIONS =========================
df['is_connected'] = (df['Talk Time Duration'] > 0).astype(int)

# Flag for RPC-related rows (before unique)
df['is_rpc_related'] = (
    df['Status'].str.contains('rpc', case=False, na=False) |
    df['Status'].str.contains('bank escalation', case=False, na=False) |
    (df['PTP Amount'] > 0)
)

# NEW: Unique RPC Count per Day per Agent using Debtor ID
# We keep only the first occurrence of each Debtor ID per (Date + CMS User)
df_unique_rpc = df[df['is_rpc_related']].drop_duplicates(subset=['Date', 'Remark By', 'Debtor ID'])

df['is_rpc'] = 0
df.loc[df_unique_rpc.index, 'is_rpc'] = 1   # Mark only unique debtors

df['is_ptp'] = (df['PTP Amount'] > 0).astype(int)
df['is_kept'] = (df['Claim Paid Amount'] > 0).astype(int)

# Outstanding Balance
df['rpc_ob']  = df['Balance'] * df['is_rpc']
df['ptp_ob']  = df['Balance'] * df['is_ptp']
df['kept_ob'] = df['Balance'] * df['is_kept']

# ========================= AGGREGATION =========================
summary = (
    df.groupby(['Date', 'Remark By'], dropna=False)
    .agg(
        Connected_Calls=('is_connected', 'sum'),
        RPC_Count=('is_rpc', 'sum'),           # Now UNIQUE by Debtor ID
        RPC_OB=('rpc_ob', 'sum'),
        PTP_Count=('is_ptp', 'sum'),
        PTP_OB=('ptp_ob', 'sum'),
        KEPT_Count=('is_kept', 'sum'),
        KEPT_OB=('kept_ob', 'sum')
    )
    .reset_index()
)

summary = summary.rename(columns={
    'Remark By': 'CMS User',
    'Connected_Calls': 'Connected Calls',
    'RPC_Count': 'RPC Count',
    'RPC_OB': 'RPC OB',
    'PTP_Count': 'PTP Count',
    'PTP_OB': 'PTP OB',
    'KEPT_Count': 'KEPT Count',
    'KEPT_OB': 'KEPT OB'
})

summary = summary[[
    'Date', 'CMS User',
    'Connected Calls', 'RPC Count', 'RPC OB',
    'PTP Count', 'PTP OB',
    'KEPT Count', 'KEPT OB'
]]

summary = summary.sort_values(by=['Date', 'CMS User'], ascending=[False, True])

# ========================= DISPLAY =========================
st.success(f"✅ Processed {len(uploaded_files)} file(s) — {len(summary):,} summary rows")

unique_dates = sorted(summary['Date'].unique(), reverse=True)

for date in unique_dates:
    st.subheader(f"📅 {date.strftime('%B %d, %Y')}")
    daily_df = summary[summary['Date'] == date].drop(columns=['Date'])
    styled = daily_df.style.set_properties(**{'text-align': 'center'})
    st.dataframe(styled, use_container_width=True, hide_index=True, height=450)

# ========================= FORMATTED EXCEL DOWNLOAD =========================
with st.sidebar:
    st.markdown("### 📥 Download Output")

    output_file = "productivity_summary_formatted.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        workbook = writer.book
        sheet = workbook.create_sheet("Summary")
        
        row_idx = 1
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        for date in unique_dates:
            date_str = date.strftime('%B %d, %Y').upper()
            
            # Merged Date Header
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            cell = sheet.cell(row=row_idx, column=1, value=date_str)
            cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            row_idx += 1
            
            # Column Headers
            headers = ['Date', 'CMS User', 'Connected Calls', 'RPC Count', 'RPC OB',
                       'PTP Count', 'PTP OB', 'KEPT Count', 'KEPT OB']
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_num, value=header)
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            row_idx += 1
            
            # Data Rows
            daily_data = summary[summary['Date'] == date]
            for _, row in daily_data.iterrows():
                for col_num, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_idx, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    if isinstance(value, (int, float)) and col_num > 2:
                        cell.number_format = '#,##0'
                row_idx += 1
            
            row_idx += 1

        # Auto-adjust widths
        for col_idx in range(1, 10):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 3, 25)
            sheet.column_dimensions[column_letter].width = adjusted_width

        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

    with open(output_file, "rb") as f:
        excel_data = f.read()

    st.download_button(
        label="📥 Download Formatted Excel File",
        data=excel_data,
        file_name=f"Productivity_Summary_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.caption("✅ RPC Count is now **unique by Debtor ID** per day per agent | All other logic preserved")